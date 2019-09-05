#reading the excel files using xlrd, openpyxl(xl, xlsx formats)
import openpyxl
import logging
import ZBase_Framework.Utilities.ConfigFile as drivepath
from ZBase_Framework.Utilities.Loghandler import Logging
import xlutils
from datetime import datetime


class Excel(Logging):

    def __init__(self,excelpath, rowinInt, colinInt):
        Logging()
        logging.info("====================Excel Test started===========================")
        self.excelpath = drivepath.Excelpath.format("Login")
        self.rowinInt = rowinInt  #
        self.colinInt = colinInt
        self.rowval = 1
        self.colval =1
        logging.info("Excel construction was created and inist")


    def excelconnection(self):
        #path = application_path
        pathexcelsheet = drivepath.Excelpath.format("Login")
        self.wb = openpyxl.load_workbook(pathexcelsheet)
        logging.info("connection was success")
        self.sh = self.wb.active
        logging.info("workssheet was active "+ self.sh.title)
        return self.wb,self.sh




    def excel_sheet(self, wbook = 'workbook',sheetname=''):
        excelpath = drivepath.Excelpath.format(wbook)
        wb = openpyxl.load_workbook(excelpath)
        logging.info("connection was success")
        sh = wb.active
        logging.info("workssheet was active "+ sh.title)
        self.sheetname = wb.get_sheet_names()
        for sheet in range(0,len(self.sheetname)):
            cmpsheet = self.sheetname[sheet]
            if str(sheetname).lower() == cmpsheet.lower() :
                    logging.info("Passed through parmeter worksheet is equal to = "+ cmpsheet)
                    sh = wb.active
                    logging.info("Sheet found :" + cmpsheet)
        row = sh.max_row
        col = sh.max_column
        return row, col,sh,wb

    def save_closeexcel(self):
        wb,sh = Excel.excelconnection(self)
        wb.save(self.excelpath)
        wb.close()







'''excelo = Excel(drivepath.Excelpath,1,2)
excelo.excelconnection()
excelo.excelwsheet("Sheet1")
excelo.excel_writeResults(2,8,"hello vinay")
excelo.save_closeexcel()'''









